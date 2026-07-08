"""Utilitas untuk mengekspor hasil kalkulasi CVI ke format Excel."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.cvi import CVIResult, ExpertRatingSummary

_STATUS_LABELS: dict[str, str] = {
    "pending": "Belum Mulai",
    "in_progress": "Sedang Berjalan",
    "completed": "Selesai",
}


def _hex_fill(hex_color: str) -> PatternFill:
    """Membuat PatternFill solid dari kode warna hex.

    Args:
        hex_color: Kode warna hex tanpa '#' (contoh: 'FFD700').

    Returns:
        Instance PatternFill dengan warna solid.
    """
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def generate_cvi_excel(result: CVIResult, expert_names: dict[str, str] | None = None) -> bytes:
    """Menghasilkan file Excel berisi hasil kalkulasi CVI.

    Struktur output:
        - Baris 1–2 : Judul instrumen dan informasi umum.
        - Baris 3   : Header tabel.
        - Baris 4+  : Data penilaian per item dengan I-CVI di kolom akhir.
        - Baris N-1 : S-CVI/Ave
        - Baris N   : S-CVI/UA

    Args:
        result: Objek CVIResult berisi hasil kalkulasi lengkap.
        expert_names: Pemetaan user_id ke nama expert (opsional, untuk header kolom).

    Returns:
        Bytes konten file Excel (.xlsx).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil CVI"

    header_fill = _hex_fill("4472C4")
    result_fill = _hex_fill("D9E1F2")
    bold_font = Font(bold=True)
    white_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Baris 1: Judul ---
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Hasil Content Validity Index — {result.instrument_name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center

    # --- Baris 2: Info umum ---
    n_valid = sum(1 for item in result.items if item.is_valid)
    ws["A2"] = "Jumlah Expert:"
    ws["B2"] = result.n_experts
    ws["C2"] = "Jumlah Item:"
    ws["D2"] = result.n_items
    ws["E2"] = f"Item Valid: {n_valid} dari {result.n_items}"
    ws["A2"].font = bold_font
    ws["C2"].font = bold_font
    ws["E2"].font = bold_font

    # --- Baris 3: Header ---
    headers = ["No", "Domain", "Item", "Jml. Relevan", "I-CVI", "Keterangan"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center

    # --- Baris 4+: Data item ---
    for row_idx, item in enumerate(result.items, start=4):
        ws.cell(row=row_idx, column=1, value=item.sequence_number).alignment = center
        ws.cell(row=row_idx, column=2, value=item.domain_name or "-")
        ws.cell(row=row_idx, column=3, value=item.content)
        ws.cell(row=row_idx, column=4, value=item.n_relevant).alignment = center
        i_cvi_cell = ws.cell(row=row_idx, column=5, value=item.i_cvi)
        i_cvi_cell.number_format = "0.00"
        i_cvi_cell.alignment = center
        status_text = "Valid" if item.is_valid else "Tidak Valid"
        status_cell = ws.cell(row=row_idx, column=6, value=status_text)
        status_cell.alignment = center
        if not item.is_valid:
            status_cell.font = Font(color="FF0000")

    # --- Baris penutup: S-CVI ---
    last_data_row = 3 + len(result.items)
    s_cvi_ave_row = last_data_row + 2
    s_cvi_ua_row = last_data_row + 3

    for row_num, label, value in [
        (s_cvi_ave_row, "S-CVI/Ave (Average Method)", result.s_cvi_ave),
        (s_cvi_ua_row, "S-CVI/UA (Universal Agreement)", result.s_cvi_ua),
    ]:
        label_cell = ws.cell(row=row_num, column=3, value=label)
        label_cell.font = bold_font
        label_cell.fill = result_fill
        value_cell = ws.cell(row=row_num, column=5, value=value)
        value_cell.number_format = "0.00"
        value_cell.font = bold_font
        value_cell.fill = result_fill
        value_cell.alignment = center

    # --- Lebar kolom ---
    column_widths = [6, 20, 60, 15, 10, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_expert_rating_excel(summary: ExpertRatingSummary, instrument_name: str) -> bytes:
    """Mengekspor penilaian satu expert untuk sebuah instrumen ke format Excel.

    Struktur output:
        - Baris 1   : Judul (nama expert + nama instrumen).
        - Baris 2   : Info umum (institusi, status, jumlah item dinilai).
        - Baris 3   : Header tabel.
        - Baris 4+  : Skor & catatan per item.

    Args:
        summary: Ringkasan penilaian seorang expert (skor + catatan per item).
        instrument_name: Nama instrumen, ditampilkan di judul dan info berkas.

    Returns:
        Isi berkas Excel (.xlsx) sebagai bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Penilaian Expert"

    header_fill = _hex_fill("4472C4")
    bold_font = Font(bold=True)
    white_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Penilaian Expert — {summary.expert_name} — {instrument_name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center

    n_rated = sum(1 for r in summary.ratings if r.relevance_score is not None)
    ws["A2"] = "Institusi:"
    ws["B2"] = summary.institution or "-"
    ws["C2"] = "Status:"
    ws["D2"] = _STATUS_LABELS.get(summary.status, summary.status)
    ws["E2"] = f"Dinilai: {n_rated}/{len(summary.ratings)}"
    ws["A2"].font = bold_font
    ws["C2"].font = bold_font
    ws["E2"].font = bold_font

    headers = ["No", "Domain", "Item", "Skor", "Relevan", "Catatan"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, rating in enumerate(summary.ratings, start=4):
        ws.cell(row=row_idx, column=1, value=rating.sequence_number).alignment = center
        ws.cell(row=row_idx, column=2, value=rating.domain_name or "-")
        ws.cell(row=row_idx, column=3, value=rating.content)
        score_value = (
            rating.relevance_score if rating.relevance_score is not None else "Belum dinilai"
        )
        ws.cell(row=row_idx, column=4, value=score_value).alignment = center
        relevan_text = (
            "Ya" if rating.is_relevant else ("Tidak" if rating.is_relevant is False else "-")
        )
        relevan_cell = ws.cell(row=row_idx, column=5, value=relevan_text)
        relevan_cell.alignment = center
        if rating.is_relevant is False:
            relevan_cell.font = Font(color="FF0000")
        ws.cell(row=row_idx, column=6, value=rating.notes or "-")

    column_widths = [6, 20, 60, 12, 10, 35]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
