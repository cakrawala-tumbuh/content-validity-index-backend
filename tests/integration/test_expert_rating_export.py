"""Integration test untuk exporter dan endpoint ekspor penilaian satu expert."""

from io import BytesIO

from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies.auth import require_admin
from app.main import app
from app.models.user import User
from app.repositories.user_repository import UserRepository
from app.schemas.cvi import ExpertRatingSummary, ItemRatingByExpert
from app.schemas.expert_assignment import AssignmentCreate
from app.schemas.instrument import InstrumentUpdate
from app.schemas.rating import RatingBulkCreate, RatingItem
from app.services.expert_assignment_service import ExpertAssignmentService
from app.services.instrument_service import InstrumentService
from app.services.item_service import ItemService
from app.services.rating_service import RatingService
from app.utils.excel_exporter import generate_expert_rating_excel
from app.utils.pdf_exporter import generate_expert_rating_pdf

from .test_rating_cvi import _setup_instrument_with_items


def _make_summary(**overrides: object) -> ExpertRatingSummary:
    """Membuat ExpertRatingSummary contoh untuk test exporter.

    Args:
        **overrides: Field yang ingin ditimpa dari nilai default.

    Returns:
        Instance ExpertRatingSummary siap dipakai test.
    """
    defaults: dict[str, object] = {
        "assignment_id": "assign-1",
        "user_id": "user-1",
        "expert_name": "Dr. Budi",
        "institution": "Universitas Test",
        "status": "completed",
        "is_active": True,
        "deadline": None,
        "ratings": [
            ItemRatingByExpert(
                item_id="item-1",
                sequence_number=1,
                content="Pertanyaan pertama",
                domain_id="dom-1",
                domain_name="Dimensi Literasi",
                relevance_score=4,
                notes=None,
                is_relevant=True,
            ),
            ItemRatingByExpert(
                item_id="item-2",
                sequence_number=2,
                content="Pertanyaan kedua",
                domain_id=None,
                domain_name=None,
                relevance_score=1,
                notes="Kurang relevan dengan konstruk",
                is_relevant=False,
            ),
        ],
    }
    defaults.update(overrides)
    return ExpertRatingSummary(**defaults)  # type: ignore[arg-type]


class TestExpertRatingExcelExporter:
    """Kumpulan test untuk generate_expert_rating_excel."""

    def test_generate_excel_menghasilkan_bytes(self) -> None:
        """generate_expert_rating_excel harus menghasilkan bytes xlsx yang valid."""
        excel_bytes = generate_expert_rating_excel(_make_summary(), "Instrumen Test")
        assert isinstance(excel_bytes, bytes)
        assert excel_bytes[:4] == b"PK\x03\x04"

    def test_generate_excel_menampilkan_domain_dan_status(self) -> None:
        """Berkas Excel harus memuat nama domain, skor, dan status assignment."""
        ws = load_workbook(
            BytesIO(generate_expert_rating_excel(_make_summary(), "Instrumen Test"))
        ).active
        assert ws["D2"].value == "Selesai"
        assert ws["B4"].value == "Dimensi Literasi"
        assert ws["D4"].value == 4
        assert ws["B5"].value == "-"
        assert ws["D5"].value == 1

    def test_generate_excel_expert_tanpa_rating(self) -> None:
        """generate_expert_rating_excel harus tetap berjalan meski belum ada rating."""
        summary = _make_summary(ratings=[])
        excel_bytes = generate_expert_rating_excel(summary, "Instrumen Kosong")
        assert len(excel_bytes) > 0


class TestExpertRatingPdfExporter:
    """Kumpulan test untuk generate_expert_rating_pdf."""

    def test_generate_pdf_menghasilkan_bytes(self) -> None:
        """generate_expert_rating_pdf harus menghasilkan bytes PDF yang valid."""
        pdf_bytes = generate_expert_rating_pdf(_make_summary(), "Instrumen Test")
        assert isinstance(pdf_bytes, bytes)
        assert pdf_bytes[:4] == b"%PDF"

    def test_generate_pdf_expert_tanpa_rating(self) -> None:
        """generate_expert_rating_pdf harus tetap berjalan meski belum ada rating."""
        summary = _make_summary(ratings=[])
        pdf_bytes = generate_expert_rating_pdf(summary, "Instrumen Kosong")
        assert len(pdf_bytes) > 0
        assert pdf_bytes[:4] == b"%PDF"

    def test_generate_pdf_content_dengan_karakter_markup(self) -> None:
        """generate_expert_rating_pdf tidak boleh gagal saat nama expert, institusi,
        content item, domain, atau catatan mengandung karakter '&', '<', '>' yang
        ditafsirkan Paragraph ReportLab sebagai mini-markup XML/HTML bila tidak di-escape.
        """
        summary = _make_summary(
            expert_name="Dr. Budi & Rekan <Ahli>",
            institution="Universitas A & B",
            ratings=[
                ItemRatingByExpert(
                    item_id="item-1",
                    sequence_number=1,
                    content="Beban kerja & tekanan (skor < 5 dianggap rendah, > 5 tinggi)",
                    domain_id="dom-1",
                    domain_name="Domain A & B",
                    relevance_score=3,
                    notes="Catatan < penting > perlu revisi & tambahan",
                    is_relevant=True,
                ),
            ],
        )
        pdf_bytes = generate_expert_rating_pdf(summary, "Instrumen & Revisi <Tim> A>B")
        assert isinstance(pdf_bytes, bytes)
        assert pdf_bytes[:4] == b"%PDF"


class TestGetExpertRatingSummary:
    """Kumpulan test untuk RatingService.get_expert_rating_summary."""

    async def test_mengembalikan_summary_milik_assignment(self, db: AsyncSession) -> None:
        """Method harus mengembalikan ExpertRatingSummary yang benar untuk assignment_id."""
        admin, expert, instrument_id = await _setup_instrument_with_items(db, "gers1")

        assign_service = ExpertAssignmentService(db)
        assignment = await assign_service.create(
            instrument_id, AssignmentCreate(user_id=expert.id), assigned_by=admin.id
        )
        item_service = ItemService(db)
        items = await item_service.get_by_instrument(instrument_id)

        rating_service = RatingService(db)
        await rating_service.bulk_submit(
            assignment.id,
            expert.id,
            RatingBulkCreate(
                ratings=[RatingItem(item_id=item.id, relevance_score=4) for item in items]
            ),
        )

        summary, instrument_name = await rating_service.get_expert_rating_summary(
            instrument_id, assignment.id
        )
        assert summary.assignment_id == assignment.id
        assert summary.expert_name == expert.full_name
        assert len(summary.ratings) == len(items)
        assert instrument_name == "Instrumen gers1"

    async def test_assignment_tidak_ditemukan_raise_404(self, db: AsyncSession) -> None:
        """Method harus melempar 404 jika assignment_id tidak ada pada instrumen ini."""
        from fastapi import HTTPException

        _, _, instrument_id = await _setup_instrument_with_items(db, "gers2")
        rating_service = RatingService(db)
        try:
            await rating_service.get_expert_rating_summary(instrument_id, "nonexistent")
            raise AssertionError("Harus melempar HTTPException")
        except HTTPException as exc:
            assert exc.status_code == 404

    async def test_instrumen_tidak_ditemukan_raise_404(self, db: AsyncSession) -> None:
        """Method harus melempar 404 jika instrumen tidak ditemukan."""
        from fastapi import HTTPException

        rating_service = RatingService(db)
        try:
            await rating_service.get_expert_rating_summary("nonexistent", "assign-x")
            raise AssertionError("Harus melempar HTTPException")
        except HTTPException as exc:
            assert exc.status_code == 404


async def _setup_assignment_with_ratings(db: AsyncSession, suffix: str) -> tuple[User, str, str]:
    """Helper menyiapkan instrumen + assignment + rating lengkap untuk test endpoint.

    Args:
        db: AsyncSession database.
        suffix: Suffix untuk ID unik.

    Returns:
        Tuple (admin, instrument_id, assignment_id).
    """
    admin, expert, instrument_id = await _setup_instrument_with_items(db, suffix)

    assign_service = ExpertAssignmentService(db)
    assignment = await assign_service.create(
        instrument_id, AssignmentCreate(user_id=expert.id), assigned_by=admin.id
    )
    item_service = ItemService(db)
    items = await item_service.get_by_instrument(instrument_id)
    rating_service = RatingService(db)
    await rating_service.bulk_submit(
        assignment.id,
        expert.id,
        RatingBulkCreate(
            ratings=[RatingItem(item_id=item.id, relevance_score=4) for item in items]
        ),
    )
    return admin, instrument_id, assignment.id


class TestExportExpertRatingExcelEndpoint:
    """Kumpulan test untuk endpoint GET .../assignments/{id}/export."""

    async def test_export_excel_mengembalikan_berkas(
        self, client: AsyncClient, db: AsyncSession
    ) -> None:
        """Endpoint harus mengembalikan 200 dengan berkas Excel."""
        admin, instrument_id, assignment_id = await _setup_assignment_with_ratings(db, "eee1")

        app.dependency_overrides[require_admin] = lambda: admin
        try:
            resp = await client.get(
                f"/api/v1/instruments/{instrument_id}/assignments/{assignment_id}/export"
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)

        assert resp.status_code == 200
        assert (
            resp.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert resp.content[:4] == b"PK\x03\x04"
        assert "attachment" in resp.headers["content-disposition"]

    async def test_export_excel_assignment_tidak_ada_raise_404(
        self, client: AsyncClient, db: AsyncSession
    ) -> None:
        """Endpoint harus mengembalikan 404 jika assignment tidak ditemukan."""
        admin, _, instrument_id = await _setup_instrument_with_items(db, "eee2")

        app.dependency_overrides[require_admin] = lambda: admin
        try:
            resp = await client.get(
                f"/api/v1/instruments/{instrument_id}/assignments/nonexistent/export"
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)

        assert resp.status_code == 404

    async def test_export_excel_instrumen_tidak_ada_raise_404(
        self, client: AsyncClient, db: AsyncSession
    ) -> None:
        """Endpoint harus mengembalikan 404 jika instrumen tidak ditemukan."""
        admin, _, _ = await _setup_instrument_with_items(db, "eee3")

        app.dependency_overrides[require_admin] = lambda: admin
        try:
            resp = await client.get(
                "/api/v1/instruments/nonexistent/assignments/nonexistent/export"
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)

        assert resp.status_code == 404


class TestExportExpertRatingPdfEndpoint:
    """Kumpulan test untuk endpoint GET .../assignments/{id}/export/pdf."""

    async def test_export_pdf_mengembalikan_berkas(
        self, client: AsyncClient, db: AsyncSession
    ) -> None:
        """Endpoint harus mengembalikan 200 dengan berkas PDF."""
        admin, instrument_id, assignment_id = await _setup_assignment_with_ratings(db, "eep1")

        app.dependency_overrides[require_admin] = lambda: admin
        try:
            resp = await client.get(
                f"/api/v1/instruments/{instrument_id}/assignments/{assignment_id}/export/pdf"
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)

        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/pdf"
        assert resp.content[:4] == b"%PDF"
        assert "attachment" in resp.headers["content-disposition"]

    async def test_export_pdf_assignment_tidak_ada_raise_404(
        self, client: AsyncClient, db: AsyncSession
    ) -> None:
        """Endpoint harus mengembalikan 404 jika assignment tidak ditemukan."""
        admin, _, instrument_id = await _setup_instrument_with_items(db, "eep2")

        app.dependency_overrides[require_admin] = lambda: admin
        try:
            resp = await client.get(
                f"/api/v1/instruments/{instrument_id}/assignments/nonexistent/export/pdf"
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)

        assert resp.status_code == 404

    async def test_export_pdf_nama_instrumen_non_latin1(
        self, client: AsyncClient, db: AsyncSession
    ) -> None:
        """Regresi: nama instrumen dengan em-dash (U+2014) tidak boleh menyebabkan 500
        karena header Content-Disposition gagal di-encode Latin-1.
        """
        admin, expert, instrument_id = await _setup_instrument_with_items(db, "eepemd")

        inst_service = InstrumentService(db)
        await inst_service.update(instrument_id, InstrumentUpdate(name="DCS — Screening"))

        assign_service = ExpertAssignmentService(db)
        assignment = await assign_service.create(
            instrument_id, AssignmentCreate(user_id=expert.id), assigned_by=admin.id
        )
        item_service = ItemService(db)
        items = await item_service.get_by_instrument(instrument_id)
        rating_service = RatingService(db)
        await rating_service.bulk_submit(
            assignment.id,
            expert.id,
            RatingBulkCreate(
                ratings=[RatingItem(item_id=item.id, relevance_score=4) for item in items]
            ),
        )

        app.dependency_overrides[require_admin] = lambda: admin
        try:
            resp = await client.get(
                f"/api/v1/instruments/{instrument_id}/assignments/{assignment.id}/export/pdf"
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)

        assert resp.status_code == 200
        resp.headers["content-disposition"].encode("latin-1")
        assert "filename*=UTF-8''" in resp.headers["content-disposition"]

    async def test_export_pdf_nama_expert_non_latin1(
        self, client: AsyncClient, db: AsyncSession
    ) -> None:
        """Regresi: nama expert (full_name) dengan em-dash tidak boleh menyebabkan 500
        karena header Content-Disposition gagal di-encode Latin-1.
        """
        admin, _, instrument_id = await _setup_instrument_with_items(db, "eepnem")

        user_repo = UserRepository(db)
        expert_em_dash = await user_repo.create(
            User(
                id="exp-eepnem-emdash",
                email="exp-eepnem-emdash@test.com",
                full_name="Dr. Budi — Ahli",
                role="expert",
                is_active=True,
            )
        )

        assign_service = ExpertAssignmentService(db)
        assignment = await assign_service.create(
            instrument_id,
            AssignmentCreate(user_id=expert_em_dash.id),
            assigned_by=admin.id,
        )
        item_service = ItemService(db)
        items = await item_service.get_by_instrument(instrument_id)
        rating_service = RatingService(db)
        await rating_service.bulk_submit(
            assignment.id,
            expert_em_dash.id,
            RatingBulkCreate(
                ratings=[RatingItem(item_id=item.id, relevance_score=4) for item in items]
            ),
        )

        app.dependency_overrides[require_admin] = lambda: admin
        try:
            resp = await client.get(
                f"/api/v1/instruments/{instrument_id}/assignments/{assignment.id}/export/pdf"
            )
        finally:
            app.dependency_overrides.pop(require_admin, None)

        assert resp.status_code == 200
        resp.headers["content-disposition"].encode("latin-1")
        assert "filename*=UTF-8''" in resp.headers["content-disposition"]
